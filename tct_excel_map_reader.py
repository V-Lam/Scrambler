import openpyxl

class TctExcelMapReader:

    def __init__(self, filename: str, workbook: str = None, worksheet: str = None):
        self.filename = filename
        self.workbook = workbook
        self.worksheet = worksheet


    def read_excel(self):
        # Open the Excel file
        '''

        # if workbook name is not specified, read using the filename
        if self.workbook is None:
            self.workbook = openpyxl.load_workbook('filename=self.filename, read_only=True, data_only=True')
        '''



        workbook = openpyxl.load_workbook('Nevada DPS Mapping FIXED.xlsx')
        worksheet = workbook.active


        # if worksheet name was not specified or found, used the active sheet instead
        # if self.worksheet is None:
            # Select the active worksheet
        # worksheet = self.workbook.active
        # else :
        #     worksheet = self.workbook[self.worksheet_name]

        # Initialize an empty dictionary
        tctMap = {}

        # Loop through each row in the worksheet
        for row in worksheet.iter_rows(values_only=True):
            # Get the value in the 2nd column (assumed to be the key)
            key = row[1]

            # Get the value in the 3rd column (assumed to be the value)
            value = row[2]

            if not key:
                # get next column
                key = row[2]
                value = row[3]

            # Check if the key and value cells are empty or not
            if key and value and value != "N/A":
                # Add the key-value pair to the dictionary
                tctMap[key] = value

        # Print the dictionary
        print(tctMap)
        return tctMap